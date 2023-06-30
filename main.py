import json
import requests
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup


def get_json_data(name):
    with open(name) as f:
        config_data = json.load(f)
    return config_data

config_data = get_json_data('config.json')

def get_products_data(page_range):
    
    if page_range <= 1:
        items_list = []
        res = requests.get(config_data['config']['product_link'])
        soup = BeautifulSoup(res.text, features="html.parser")
        products = soup.select('a[class="goods-tile__heading \
                               ng-star-inserted"]')
        product_price = soup.select(
            'span[class="goods-tile__price-value"]'
            )
        
        for product in products:
            href = product['href']
            title = product['title']
            response = requests.get(r"%s" % (href))
            product_soup = BeautifulSoup(response.text, 
                                         features='html.parser')
            if product_soup.find("p", attrs={
                "class":"product-price__big \
                product-price__big-color-red"}) is not None:
                price = product_soup.find("p", attrs={
                    "class":"product-price__big \
                    product-price__big-color-red"}).text
                items_list.append({'title': title,
                                'link': href,
                                'price': price
                                })
        
        return items_list
    
    else:
        items_list = []
        url = config_data['config']['product_link']
        url_elements = url.split('/')
        url_elements.pop(-1)
        new_url = ''
        
        for i in url_elements:
            new_url += i + '/'
        
        for i in range(0, page_range):
            page_url = new_url + f"page={int(i)}/"
            res = requests.get(page_url)
            soup = BeautifulSoup(res.text, features="html.parser")
            products = soup.select('a[class="goods-tile__heading \
                                   ng-star-inserted"]')
        
            for product in products:
                href = product['href']
                title = product['title']
                response = requests.get(r"%s" % (href))
                product_soup = BeautifulSoup(response.text, features='html.parser')
                
                if product_soup.find("p", attrs={
                    "class":"product-price__big \
                    product-price__big-color-red"}) is not None:
                    price = product_soup.find("p", attrs={
                        "class":"product-price__big \
                        product-price__big-color-red"}).text
                    items_list.append({'title': title,
                                    'link': href,
                                    'price': price
                                    })
                
        return items_list
            
        
def write_data_to_exel(data: list):
    wb = Workbook()
    ws = wb.active
    
    for i in range(len(data)):
        ws[f'A{str(int(i) + 1)}'] = data[i]['title']
        ws[f'B{str(int(i) + 1)}'] = data[i]['link']
        ws[f'C{str(int(i) + 1)}'] = data[i]['price']
        
    wb.save(config_data['config']['exel_file_name'])
            
def main():
    data = get_products_data(5)
    write_data_to_exel(data)
    
if __name__ == "__main__":
    main()
    